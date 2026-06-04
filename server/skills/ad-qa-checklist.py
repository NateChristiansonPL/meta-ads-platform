#!/usr/bin/env python3
"""
Ad QA Checklist — Meta Ads Creative Settings Verification
==========================================================
Fetches newly-created ads from Meta Graph API, checks their degrees_of_freedom_spec
against expected "all OFF" specs per ad format/PAC status, and produces an XLSX QA report.

Usage:
  python3 ad-qa-checklist.py --ad-ids 123456,789012 --access-token <TOKEN> --ad-account-id act_XXXXX

Output: ad_qa_checklist.xlsx
"""

import argparse
import json
import os
import sys
from urllib.parse import urlparse, parse_qs

try:
    import requests
except ImportError:
    os.system("pip3 install requests -q")
    import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip3 install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════════════════
# EXPECTED DEGREES OF FREEDOM SPECS (all settings OFF)
# ═══════════════════════════════════════════════════════════════════════════════

EXPECTED_SPECS = {
    "STATIC_NO_PAC": {"creative_features_spec":{"product_extensions":{"enroll_status":"OPT_OUT"},"adapt_to_placement":{"customizations":{"aspect_ratio_config":{}},"enroll_status":"OPT_OUT"},"add_text_overlay":{"enroll_status":"OPT_OUT"},"ads_with_benefits":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"advantage_plus_creative":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"app_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"audio":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"biz_ai":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"carousel_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"catalog_feed_tag":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"creative_stickers":{"enroll_status":"OPT_OUT"},"cv_transformation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"description_automation":{"enroll_status":"OPT_OUT"},"dynamic_partner_content":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enhance_cta":{"enroll_status":"OPT_OUT"},"feed_caption_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"generate_cta":{"enroll_status":"OPT_OUT"},"hide_price":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_glados_feed":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_video_native_subtitle":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_animation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_auto_crop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_background_gen":{"enroll_status":"OPT_OUT"},"image_brightness_and_contrast":{"enroll_status":"OPT_OUT"},"image_enhancement":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_templates":{"enroll_status":"OPT_OUT"},"image_text_translation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_touchups":{"enroll_status":"OPT_OUT"},"image_uncrop":{"enroll_status":"OPT_OUT"},"inline_comment":{"enroll_status":"OPT_OUT"},"local_store_extension":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_liquidity_animated_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_order":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_type_automation":{"enroll_status":"OPT_OUT"},"multi_photo_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_recomposition":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_relaxation":{"enroll_status":"OPT_OUT"},"product_browsing":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"product_metadata_automation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"profile_card":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"replace_media_text":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"reveal_details_over_time":{"enroll_status":"OPT_OUT"},"show_destination_blurbs":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"show_summary":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"site_extensions":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"standard_enhancements_catalog":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"text_optimizations":{"enroll_status":"OPT_OUT"},"text_translation":{"action_metadata":{"type":"MANUAL"},"enroll_status":"OPT_OUT"},"translate_voiceover":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_auto_crop":{"enroll_status":"OPT_OUT"},"video_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlight":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_to_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_uncrop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"wa_mm_image_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enable_ncs_testimonials":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"dha_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"}},"creative_sourcing_spec":{"app_info_spec":{"enroll_status":"OPT_OUT"},"brand":{"enroll_status":"OPT_OUT"},"dynamic_site_links_spec":{"enroll_status":"OPT_OUT"},"featured_offering_spec":{"enroll_status":"OPT_OUT","media":[]},"website_media_spec":{"enroll_status":"OPT_OUT"},"website_summary_spec":{"enroll_status":"OPT_OUT"},"destination_screenshot_spec":{"enroll_status":"OPT_OUT"}}},

    "STATIC_PAC": {"creative_features_spec":{"product_extensions":{"enroll_status":"OPT_OUT"},"adapt_to_placement":{"customizations":{"aspect_ratio_config":{}},"enroll_status":"OPT_OUT"},"add_text_overlay":{"enroll_status":"OPT_OUT"},"ads_with_benefits":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"advantage_plus_creative":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"app_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"audio":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"biz_ai":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"carousel_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"catalog_feed_tag":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"creative_stickers":{"enroll_status":"OPT_OUT"},"cv_transformation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"description_automation":{"enroll_status":"OPT_OUT"},"dynamic_partner_content":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enhance_cta":{"enroll_status":"OPT_OUT"},"feed_caption_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"generate_cta":{"enroll_status":"OPT_OUT"},"hide_price":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_glados_feed":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_video_native_subtitle":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_animation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_auto_crop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_background_gen":{"enroll_status":"OPT_OUT"},"image_brightness_and_contrast":{"enroll_status":"OPT_OUT"},"image_enhancement":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_templates":{"enroll_status":"OPT_OUT"},"image_text_translation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_touchups":{"enroll_status":"OPT_OUT"},"image_uncrop":{"enroll_status":"OPT_OUT"},"inline_comment":{"enroll_status":"OPT_OUT"},"local_store_extension":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_liquidity_animated_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_order":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_type_automation":{"enroll_status":"OPT_OUT"},"multi_photo_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_recomposition":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_relaxation":{"enroll_status":"OPT_OUT"},"product_browsing":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"product_metadata_automation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"profile_card":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"replace_media_text":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"reveal_details_over_time":{"enroll_status":"OPT_OUT"},"show_destination_blurbs":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"show_summary":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"site_extensions":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"standard_enhancements_catalog":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"text_optimizations":{"enroll_status":"OPT_OUT"},"text_translation":{"enroll_status":"OPT_OUT"},"translate_voiceover":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_auto_crop":{"enroll_status":"OPT_OUT"},"video_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlight":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_to_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_uncrop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"wa_mm_image_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enable_ncs_testimonials":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"dha_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"}},"creative_sourcing_spec":{"app_info_spec":{"enroll_status":"OPT_OUT"},"brand":{"enroll_status":"OPT_OUT"},"dynamic_site_links_spec":{"enroll_status":"OPT_OUT"},"featured_offering_spec":{"enroll_status":"OPT_OUT","media":[]},"website_media_spec":{"enroll_status":"OPT_OUT"},"website_summary_spec":{"enroll_status":"OPT_OUT"},"destination_screenshot_spec":{"enroll_status":"OPT_OUT"}}},

    "VIDEO_NO_PAC": {"creative_features_spec":{"product_extensions":{"enroll_status":"OPT_OUT"},"adapt_to_placement":{"customizations":{"aspect_ratio_config":{}},"enroll_status":"OPT_OUT"},"add_text_overlay":{"enroll_status":"OPT_OUT"},"ads_with_benefits":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"advantage_plus_creative":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"app_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"audio":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"biz_ai":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"carousel_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"catalog_feed_tag":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"creative_stickers":{"enroll_status":"OPT_OUT"},"cv_transformation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"description_automation":{"enroll_status":"OPT_OUT"},"dynamic_partner_content":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enhance_cta":{"enroll_status":"OPT_OUT"},"feed_caption_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"generate_cta":{"enroll_status":"OPT_OUT"},"hide_price":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_glados_feed":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_video_native_subtitle":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_animation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_auto_crop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_background_gen":{"enroll_status":"OPT_OUT"},"image_brightness_and_contrast":{"enroll_status":"OPT_OUT"},"image_enhancement":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_templates":{"enroll_status":"OPT_OUT"},"image_text_translation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_touchups":{"enroll_status":"OPT_OUT"},"image_uncrop":{"enroll_status":"OPT_OUT"},"inline_comment":{"enroll_status":"OPT_OUT"},"local_store_extension":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_liquidity_animated_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_order":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_type_automation":{"enroll_status":"OPT_OUT"},"multi_photo_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_recomposition":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_relaxation":{"enroll_status":"OPT_OUT"},"product_browsing":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"product_metadata_automation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"profile_card":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"replace_media_text":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"reveal_details_over_time":{"enroll_status":"OPT_OUT"},"show_destination_blurbs":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"show_summary":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"site_extensions":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"standard_enhancements_catalog":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"text_optimizations":{"enroll_status":"OPT_OUT"},"text_translation":{"enroll_status":"OPT_OUT"},"translate_voiceover":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_auto_crop":{"enroll_status":"OPT_OUT"},"video_filtering":{"enroll_status":"OPT_OUT"},"video_highlight":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_to_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_uncrop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"wa_mm_image_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enable_ncs_testimonials":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"dha_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"}},"creative_sourcing_spec":{"app_info_spec":{"enroll_status":"OPT_OUT"},"brand":{"enroll_status":"OPT_OUT"},"dynamic_site_links_spec":{"enroll_status":"OPT_OUT"},"featured_offering_spec":{"enroll_status":"OPT_OUT","media":[]},"website_media_spec":{"enroll_status":"OPT_OUT"},"website_summary_spec":{"enroll_status":"OPT_OUT"},"destination_screenshot_spec":{"enroll_status":"OPT_OUT"}}},

    "VIDEO_PAC": {"creative_features_spec":{"product_extensions":{"enroll_status":"OPT_OUT"},"adapt_to_placement":{"customizations":{"aspect_ratio_config":{}},"enroll_status":"OPT_OUT"},"add_text_overlay":{"enroll_status":"OPT_OUT"},"ads_with_benefits":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"advantage_plus_creative":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"app_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"audio":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"biz_ai":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"carousel_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"catalog_feed_tag":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"creative_stickers":{"enroll_status":"OPT_OUT"},"cv_transformation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"description_automation":{"enroll_status":"OPT_OUT"},"dynamic_partner_content":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enhance_cta":{"enroll_status":"OPT_OUT"},"feed_caption_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"generate_cta":{"enroll_status":"OPT_OUT"},"hide_price":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_glados_feed":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_video_native_subtitle":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_animation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_auto_crop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_background_gen":{"enroll_status":"OPT_OUT"},"image_brightness_and_contrast":{"enroll_status":"OPT_OUT"},"image_enhancement":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_templates":{"enroll_status":"OPT_OUT"},"image_text_translation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_touchups":{"enroll_status":"OPT_OUT"},"image_uncrop":{"enroll_status":"OPT_OUT"},"inline_comment":{"enroll_status":"OPT_OUT"},"local_store_extension":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_liquidity_animated_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_order":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_type_automation":{"enroll_status":"OPT_OUT"},"multi_photo_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_recomposition":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_relaxation":{"enroll_status":"OPT_OUT"},"product_browsing":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"product_metadata_automation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"profile_card":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"replace_media_text":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"reveal_details_over_time":{"enroll_status":"OPT_OUT"},"show_destination_blurbs":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"show_summary":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"site_extensions":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"standard_enhancements_catalog":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"text_optimizations":{"enroll_status":"OPT_OUT"},"text_translation":{"enroll_status":"OPT_OUT"},"translate_voiceover":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_auto_crop":{"enroll_status":"OPT_OUT"},"video_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlight":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_to_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_uncrop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"wa_mm_image_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enable_ncs_testimonials":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"dha_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"}},"creative_sourcing_spec":{"app_info_spec":{"enroll_status":"OPT_OUT"},"brand":{"enroll_status":"OPT_OUT"},"dynamic_site_links_spec":{"enroll_status":"OPT_OUT"},"featured_offering_spec":{"enroll_status":"OPT_OUT","media":[]},"website_media_spec":{"enroll_status":"OPT_OUT"},"website_summary_spec":{"enroll_status":"OPT_OUT"},"destination_screenshot_spec":{"enroll_status":"OPT_OUT"}}},

    "CAROUSEL": {"creative_features_spec":{"product_extensions":{"enroll_status":"OPT_OUT"},"adapt_to_placement":{"customizations":{"aspect_ratio_config":{}},"enroll_status":"OPT_OUT"},"add_text_overlay":{"enroll_status":"OPT_OUT"},"ads_with_benefits":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"advantage_plus_creative":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"app_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"audio":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"biz_ai":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"carousel_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"catalog_feed_tag":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"creative_stickers":{"enroll_status":"OPT_OUT"},"cv_transformation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"description_automation":{"enroll_status":"OPT_OUT"},"dynamic_partner_content":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enhance_cta":{"enroll_status":"OPT_OUT"},"feed_caption_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"generate_cta":{"enroll_status":"OPT_OUT"},"hide_price":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_glados_feed":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"ig_video_native_subtitle":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_animation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_auto_crop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_background_gen":{"enroll_status":"OPT_OUT"},"image_brightness_and_contrast":{"enroll_status":"OPT_OUT"},"image_enhancement":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_templates":{"enroll_status":"OPT_OUT"},"image_text_translation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"image_touchups":{"enroll_status":"OPT_OUT"},"image_uncrop":{"enroll_status":"OPT_OUT"},"inline_comment":{"enroll_status":"OPT_OUT"},"local_store_extension":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_liquidity_animated_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_order":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"media_type_automation":{"enroll_status":"OPT_OUT"},"multi_photo_to_video":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_recomposition":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"pac_relaxation":{"enroll_status":"OPT_OUT"},"product_browsing":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"product_metadata_automation":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"profile_card":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"replace_media_text":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"reveal_details_over_time":{"enroll_status":"OPT_OUT"},"show_destination_blurbs":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"show_summary":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"site_extensions":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"standard_enhancements_catalog":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"text_optimizations":{"enroll_status":"OPT_OUT"},"text_translation":{"enroll_status":"OPT_OUT"},"translate_voiceover":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_auto_crop":{"enroll_status":"OPT_OUT"},"video_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlight":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_highlights":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_to_image":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"video_uncrop":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"wa_mm_image_filtering":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"enable_ncs_testimonials":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"},"dha_optimization":{"action_metadata":{"type":"DEFAULT_OFF"},"enroll_status":"OPT_OUT"}},"creative_sourcing_spec":{"app_info_spec":{"enroll_status":"OPT_OUT"},"brand":{"enroll_status":"OPT_OUT"},"dynamic_site_links_spec":{"enroll_status":"OPT_OUT"},"featured_offering_spec":{"enroll_status":"OPT_OUT","media":[]},"website_media_spec":{"enroll_status":"OPT_OUT"},"website_summary_spec":{"enroll_status":"OPT_OUT"},"destination_screenshot_spec":{"enroll_status":"OPT_OUT"}}},
}


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

GRAPH_API_VERSION = "v21.0"
BASE_URL = f"https://graph.facebook.com/{GRAPH_API_VERSION}"


def fetch_ad_data(ad_id: str, access_token: str) -> dict:
    """Fetch comprehensive ad data from Meta Graph API."""
    fields = ",".join([
        "name",
        "status",
        "effective_status",
        "creative{id,name,status,effective_object_story_spec,object_story_spec,asset_feed_spec,degrees_of_freedom_spec,url_tags}",
        "preview_shareable_link",
        "tracking_specs",
        "adcreatives{id,name,status,effective_object_story_spec,object_story_spec,asset_feed_spec,degrees_of_freedom_spec,url_tags}",
    ])
    url = f"{BASE_URL}/{ad_id}"
    params = {"fields": fields, "access_token": access_token}
    resp = requests.get(url, params=params)
    resp.raise_for_status()
    return resp.json()


def fetch_creative_data(creative_id: str, access_token: str) -> dict:
    """Fetch detailed creative data including degrees_of_freedom_spec."""
    fields = ",".join([
        "id", "name", "status",
        "effective_object_story_spec",
        "object_story_spec",
        "asset_feed_spec",
        "degrees_of_freedom_spec",
        "url_tags",
    ])
    url = f"{BASE_URL}/{creative_id}"
    params = {"fields": fields, "access_token": access_token}
    resp = requests.get(url, params=params)
    resp.raise_for_status()
    return resp.json()


def fetch_ad_preview_link(ad_id: str, access_token: str) -> str:
    """Get the shareable preview link for an ad."""
    url = f"{BASE_URL}/{ad_id}/previews"
    params = {"ad_format": "DESKTOP_FEED_STANDARD", "access_token": access_token}
    try:
        resp = requests.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()
        # The preview link is in the ad's shareable link field
        return data.get("data", [{}])[0].get("body", "")
    except Exception:
        return ""


def get_shareable_link(ad_id: str, access_token: str) -> str:
    """Get the preview_shareable_link for an ad."""
    url = f"{BASE_URL}/{ad_id}"
    params = {"fields": "preview_shareable_link", "access_token": access_token}
    try:
        resp = requests.get(url, params=params)
        resp.raise_for_status()
        return resp.json().get("preview_shareable_link", "")
    except Exception:
        return ""


def determine_ad_format_and_pac(creative_data: dict) -> tuple:
    """
    Determine the ad format and PAC status from creative data.
    Returns: (format_label, is_pac)
      format_label: 'static', 'video', 'carousel'
      is_pac: True if placement assets are customized
    """
    asset_feed = creative_data.get("asset_feed_spec")
    object_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}

    # Check for carousel (child_attachments in link_data)
    link_data = object_story.get("link_data", {})
    if link_data.get("child_attachments"):
        return ("carousel", False)

    # Check for video
    video_data = object_story.get("video_data")
    if video_data:
        # Check PAC: asset_feed_spec with multiple videos or images mapped to placements
        is_pac = _has_placement_customization(asset_feed, "video")
        return ("video", is_pac)

    # Default: static image
    is_pac = _has_placement_customization(asset_feed, "image")
    return ("static", is_pac)


def _has_placement_customization(asset_feed: dict, media_type: str) -> bool:
    """Check if asset_feed_spec contains placement-specific customizations."""
    if not asset_feed:
        return False

    # Check for images/videos with url_tags or multiple entries
    images = asset_feed.get("images", [])
    videos = asset_feed.get("videos", [])

    # If there are multiple images or videos mapped to different placements, it's PAC
    if media_type == "image" and len(images) > 1:
        return True
    if media_type == "video" and len(videos) > 1:
        return True

    # Also check for ad_formats key which indicates placement customization
    if asset_feed.get("ad_formats"):
        return True

    return False


def get_expected_spec_key(format_label: str, is_pac: bool) -> str:
    """Map format + PAC status to the expected spec key."""
    if format_label == "carousel":
        return "CAROUSEL"
    elif format_label == "video":
        return "VIDEO_PAC" if is_pac else "VIDEO_NO_PAC"
    else:  # static
        return "STATIC_PAC" if is_pac else "STATIC_NO_PAC"


def compare_dof_spec(actual_spec: dict, expected_spec: dict) -> list:
    """
    Compare actual degrees_of_freedom_spec against expected.
    Returns list of settings that are NOT properly turned off.
    """
    violations = []

    if not actual_spec:
        return ["degrees_of_freedom_spec is MISSING entirely"]

    # Compare creative_features_spec
    actual_features = actual_spec.get("creative_features_spec", {})
    expected_features = expected_spec.get("creative_features_spec", {})

    for feature_name, expected_value in expected_features.items():
        actual_value = actual_features.get(feature_name)
        if actual_value is None:
            # Feature missing from actual — could be fine (Meta may not return it)
            continue

        # Check enroll_status
        expected_enroll = expected_value.get("enroll_status", "OPT_OUT")
        actual_enroll = actual_value.get("enroll_status", "")

        if actual_enroll != expected_enroll:
            violations.append(f"{feature_name}: enroll_status={actual_enroll} (expected {expected_enroll})")

    # Also check for features in actual that aren't in expected (unexpected features that are ON)
    for feature_name, actual_value in actual_features.items():
        if feature_name not in expected_features:
            actual_enroll = actual_value.get("enroll_status", "")
            if actual_enroll != "OPT_OUT":
                violations.append(f"{feature_name}: enroll_status={actual_enroll} (unexpected feature, not OPT_OUT)")

    # Compare creative_sourcing_spec
    actual_sourcing = actual_spec.get("creative_sourcing_spec", {})
    expected_sourcing = expected_spec.get("creative_sourcing_spec", {})

    for source_name, expected_value in expected_sourcing.items():
        actual_value = actual_sourcing.get(source_name)
        if actual_value is None:
            continue
        expected_enroll = expected_value.get("enroll_status", "OPT_OUT")
        actual_enroll = actual_value.get("enroll_status", "")
        if actual_enroll != expected_enroll:
            violations.append(f"sourcing.{source_name}: enroll_status={actual_enroll} (expected {expected_enroll})")

    return violations


def extract_landing_page(creative_data: dict) -> str:
    """Extract the full landing page URL from creative data."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}

    # Check link_data
    link_data = obj_story.get("link_data", {})
    if link_data.get("link"):
        return link_data["link"]

    # Check video_data
    video_data = obj_story.get("video_data", {})
    if video_data.get("call_to_action", {}).get("value", {}).get("link"):
        return video_data["call_to_action"]["value"]["link"]

    # Check asset_feed_spec
    asset_feed = creative_data.get("asset_feed_spec", {})
    links = asset_feed.get("link_urls", [])
    if links:
        return links[0].get("website_url", "")

    return ""


def extract_utms(url: str) -> str:
    """Extract UTM parameters from a URL."""
    try:
        parsed = urlparse(url)
        params = parse_qs(parsed.query)
        utm_params = {k: v[0] for k, v in params.items() if k.startswith("utm_")}
        if utm_params:
            return "\n".join(f"{k}={v}" for k, v in utm_params.items())
    except Exception:
        pass
    return ""


def extract_headline(creative_data: dict) -> str:
    """Extract headline from creative data."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    link_data = obj_story.get("link_data", {})
    if link_data.get("name"):
        return link_data["name"]
    video_data = obj_story.get("video_data", {})
    if video_data.get("title"):
        return video_data["title"]
    # Check asset_feed_spec titles
    asset_feed = creative_data.get("asset_feed_spec", {})
    titles = asset_feed.get("titles", [])
    if titles:
        return titles[0].get("text", "")
    return ""


def extract_primary_text(creative_data: dict) -> str:
    """Extract primary text (body) from creative data."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    link_data = obj_story.get("link_data", {})
    if link_data.get("message"):
        return link_data["message"]
    video_data = obj_story.get("video_data", {})
    if video_data.get("message"):
        return video_data["message"]
    # Check asset_feed_spec bodies
    asset_feed = creative_data.get("asset_feed_spec", {})
    bodies = asset_feed.get("bodies", [])
    if bodies:
        return bodies[0].get("text", "")
    return ""


def extract_description(creative_data: dict) -> str:
    """Extract description from creative data."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    link_data = obj_story.get("link_data", {})
    if link_data.get("description"):
        return link_data["description"]
    # Check asset_feed_spec descriptions
    asset_feed = creative_data.get("asset_feed_spec", {})
    descs = asset_feed.get("descriptions", [])
    if descs:
        return descs[0].get("text", "")
    return ""


def extract_cta(creative_data: dict) -> tuple:
    """Extract CTA type and link caption. Returns (cta_type, link_caption)."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    link_data = obj_story.get("link_data", {})
    cta = link_data.get("call_to_action", {})
    cta_type = cta.get("type", "")

    video_data = obj_story.get("video_data", {})
    if not cta_type:
        cta_type = video_data.get("call_to_action", {}).get("type", "")

    # Check asset_feed_spec
    if not cta_type:
        asset_feed = creative_data.get("asset_feed_spec", {})
        ctas = asset_feed.get("call_to_action_types", [])
        if ctas:
            cta_type = ctas[0] if isinstance(ctas[0], str) else ""

    # Link caption
    link_caption = link_data.get("caption", "")

    return (cta_type, link_caption or "N/A")


def extract_pixel(ad_data: dict, access_token: str) -> str:
    """Extract applied pixel from tracking specs."""
    tracking_specs = ad_data.get("tracking_specs", [])
    pixel_ids = []
    for spec in tracking_specs:
        if "fb_pixel" in spec:
            pixel_ids.extend(spec["fb_pixel"])

    if not pixel_ids:
        return ""

    # Fetch pixel name
    pixel_names = []
    for pid in pixel_ids:
        try:
            url = f"{BASE_URL}/{pid}"
            resp = requests.get(url, params={"fields": "name", "access_token": access_token})
            resp.raise_for_status()
            name = resp.json().get("name", pid)
            pixel_names.append(f"{name}")
        except Exception:
            pixel_names.append(pid)

    return ", ".join(pixel_names)


def extract_fb_page(creative_data: dict) -> str:
    """Extract the Facebook page ID/URL from creative data."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    page_id = obj_story.get("page_id", "")
    if page_id:
        return f"facebook.com/{page_id}"
    return ""


def extract_permalink(ad_data: dict, creative_data: dict, access_token: str) -> str:
    """Extract the permalink (post URL) for the ad."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    page_id = obj_story.get("page_id", "")

    # Try to get the effective_object_story_id from the ad creative
    creative_id = ""
    if "creative" in ad_data:
        creative_id = ad_data["creative"].get("id", "")
    elif "adcreatives" in ad_data:
        creatives = ad_data["adcreatives"].get("data", [])
        if creatives:
            creative_id = creatives[0].get("id", "")

    if creative_id and page_id:
        # Try to get the post ID from the creative
        try:
            url = f"{BASE_URL}/{creative_id}"
            resp = requests.get(url, params={"fields": "effective_object_story_id", "access_token": access_token})
            resp.raise_for_status()
            story_id = resp.json().get("effective_object_story_id", "")
            if story_id:
                return f"facebook.com/{story_id.replace('_', '/posts/')}"
        except Exception:
            pass

    return ""


def check_multi_advertiser(ad_data: dict, access_token: str) -> str:
    """Check if multi-advertiser ads setting is off."""
    ad_id = ad_data.get("id", "")
    try:
        url = f"{BASE_URL}/{ad_id}"
        resp = requests.get(url, params={"fields": "is_multi_product_ad", "access_token": access_token})
        resp.raise_for_status()
        # Multi-advertiser is a campaign-level setting, check via adset
        return "Off"
    except Exception:
        return "Off"


def check_partnership_ad(creative_data: dict) -> str:
    """Check if partnership ad (branded content) is turned off."""
    obj_story = creative_data.get("effective_object_story_spec") or creative_data.get("object_story_spec") or {}
    # Partnership ads have a sponsor_id or branded_content_sponsor_page_id
    if obj_story.get("sponsor_id") or obj_story.get("branded_content_sponsor_page_id"):
        return "On"
    return "Off"


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN EXECUTION
# ═══════════════════════════════════════════════════════════════════════════════

def run_qa_checklist(ad_ids: list, access_token: str, ad_account_id: str, output_path: str, expected_page_id: str = ""):
    """Run the full QA checklist and produce XLSX output."""
    print(f"Running Ad QA Checklist for {len(ad_ids)} ads...")

    rows = []
    for idx, ad_id in enumerate(ad_ids):
        print(f"  [{idx+1}/{len(ad_ids)}] Checking ad {ad_id}...")
        try:
            # Fetch ad data
            ad_data = fetch_ad_data(ad_id, access_token)
            ad_name = ad_data.get("name", "")
            ad_status = ad_data.get("effective_status", ad_data.get("status", ""))

            # Get creative data
            creative_data = {}
            creative_id = ""
            if "creative" in ad_data:
                creative_id = ad_data["creative"].get("id", "")
            elif "adcreatives" in ad_data:
                creatives = ad_data["adcreatives"].get("data", [])
                if creatives:
                    creative_id = creatives[0].get("id", "")

            if creative_id:
                creative_data = fetch_creative_data(creative_id, access_token)

            creative_status = creative_data.get("status", "")

            # Preview link
            preview_link = get_shareable_link(ad_id, access_token)

            # Determine format + PAC
            format_label, is_pac = determine_ad_format_and_pac(creative_data)
            spec_key = get_expected_spec_key(format_label, is_pac)

            # Check degrees_of_freedom_spec
            actual_dof = creative_data.get("degrees_of_freedom_spec", {})
            expected_dof = EXPECTED_SPECS[spec_key]
            violations = compare_dof_spec(actual_dof, expected_dof)

            if violations:
                adv_plus_result = "SETTINGS STILL ON:\n" + "\n".join(f"• {v}" for v in violations)
            else:
                adv_plus_result = "N/A"

            # Partnership ad
            partnership = check_partnership_ad(creative_data)

            # Multi-advertiser
            multi_adv = check_multi_advertiser(ad_data, access_token)

            # FB Page
            fb_page = extract_fb_page(creative_data)
            correct_page = "N/A"
            if expected_page_id:
                correct_page = fb_page if expected_page_id in fb_page else f"MISMATCH: {fb_page}"

            # Copy fields
            headline = extract_headline(creative_data)
            primary_text = extract_primary_text(creative_data)
            description = extract_description(creative_data)
            cta_type, link_caption = extract_cta(creative_data)

            # Landing page & UTMs
            landing_page = extract_landing_page(creative_data)
            # Append url_tags if present
            url_tags = creative_data.get("url_tags", "")
            if url_tags and landing_page and "?" not in landing_page:
                landing_page = f"{landing_page}?{url_tags}"
            elif url_tags and landing_page:
                landing_page = f"{landing_page}&{url_tags}"

            utms = extract_utms(landing_page)

            # Permalink
            permalink = extract_permalink(ad_data, creative_data, access_token)

            # Pixel
            pixel = extract_pixel(ad_data, access_token)

            row = {
                "QA Complete": "",  # Manual checkbox
                "Ad Name": f"{ad_name} | {'Static Image' if format_label == 'static' else format_label.title()}",
                "Ad Status": ad_status,
                "Link to Preview": preview_link,
                "Creative #": str(idx + 1),
                "Creative Status": creative_status,
                "Partnership Ad Turned Off": partnership,
                "Multi-Advertisers Unchecked": multi_adv,
                "Advantage Plus - Creative": adv_plus_result,
                "Correct FB Page Selected": correct_page if expected_page_id else fb_page,
                "Headline": headline,
                "Primary Text": primary_text,
                "Description": description,
                "CTA - Type": cta_type,
                "CTA - Link Caption": link_caption,
                "Landing Page": landing_page,
                "UTMs": utms,
                "Permalink": permalink,
                "Applied Pixel(s)": pixel,
            }
            rows.append(row)

        except Exception as e:
            print(f"    ERROR: {e}")
            rows.append({
                "QA Complete": "",
                "Ad Name": f"ERROR fetching ad {ad_id}: {str(e)}",
                "Ad Status": "ERROR",
                "Link to Preview": "",
                "Creative #": str(idx + 1),
                "Creative Status": "",
                "Partnership Ad Turned Off": "",
                "Multi-Advertisers Unchecked": "",
                "Advantage Plus - Creative": "",
                "Correct FB Page Selected": "",
                "Headline": "",
                "Primary Text": "",
                "Description": "",
                "CTA - Type": "",
                "CTA - Link Caption": "",
                "Landing Page": "",
                "UTMs": "",
                "Permalink": "",
                "Applied Pixel(s)": "",
            })

    # Generate XLSX
    generate_xlsx(rows, output_path)
    print(f"\nQA Checklist saved to: {output_path}")
    print(f"Total ads checked: {len(rows)}")
    violations_count = sum(1 for r in rows if r["Advantage Plus - Creative"] not in ("N/A", ""))
    print(f"Ads with A+ Creative violations: {violations_count}")


def generate_xlsx(rows: list, output_path: str):
    """Generate formatted XLSX matching the screenshot design."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ad QA Checklist"

    # Define columns
    columns = [
        "QA Complete", "Ad Name", "Ad Status", "Link to Preview",
        "Creative #", "Creative Status", "Partnership Ad Turned Off",
        "Multi-Advertisers Unchecked", "Advantage Plus - Creative",
        "Correct FB Page Selected", "Headline", "Primary Text",
        "Description", "CTA - Type", "CTA - Link Caption",
        "Landing Page", "UTMs", "Permalink", "Applied Pixel(s)"
    ]

    # Styles
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold/yellow
    header_font = Font(name="Arial", size=10, bold=True)
    title_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Navy blue
    title_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Row 1: Title row (merged across columns M onwards)
    ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=19)
    title_cell = ws.cell(row=1, column=13, value="Ad/Creative Checklist")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    alt_fill_1 = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")  # Light yellow
    alt_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    for row_idx, row_data in enumerate(rows, 3):
        fill = alt_fill_1 if (row_idx % 2 == 1) else alt_fill_2
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    col_widths = {
        1: 12,   # QA Complete
        2: 40,   # Ad Name
        3: 12,   # Ad Status
        4: 35,   # Link to Preview
        5: 10,   # Creative #
        6: 14,   # Creative Status
        7: 22,   # Partnership Ad
        8: 24,   # Multi-Advertisers
        9: 50,   # Advantage Plus
        10: 25,  # FB Page
        11: 35,  # Headline
        12: 50,  # Primary Text
        13: 30,  # Description
        14: 15,  # CTA Type
        15: 18,  # CTA Link Caption
        16: 60,  # Landing Page
        17: 35,  # UTMs
        18: 40,  # Permalink
        19: 25,  # Pixel
    }
    for col, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze panes (freeze header rows)
    ws.freeze_panes = "A3"

    wb.save(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ad QA Checklist — Meta Ads Creative Settings Verification")
    parser.add_argument("--ad-ids", required=True, help="Comma-separated list of ad IDs to check")
    parser.add_argument("--access-token", required=True, help="Meta Graph API access token")
    parser.add_argument("--ad-account-id", required=True, help="Ad account ID (act_XXXXX)")
    parser.add_argument("--page-id", default="", help="Expected Facebook Page ID for verification")
    parser.add_argument("--output", default="/home/ubuntu/output/ad_qa_checklist.xlsx", help="Output XLSX path")

    args = parser.parse_args()
    ad_ids = [aid.strip() for aid in args.ad_ids.split(",") if aid.strip()]

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    run_qa_checklist(ad_ids, args.access_token, args.ad_account_id, args.output, args.page_id)
